import os
import mysql.connector
from mysql.connector import connect, Error
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px
from openpyxl import load_workbook

os.makedirs("charts", exist_ok=True)
os.makedirs("exports", exist_ok=True)

try:
    with connect(
        host="localhost",
        user=input("Имя пользователя: "),
        password=input("Пароль: "),
        database="games",
    ) as connection:
        print("Успешное подключение к базе!")

        def make_chart(df, chart_type, title, x=None, y=None, hue=None, filename="chart.png"):
            plt.figure(figsize=(8, 5))
            if chart_type == "pie":
                df.set_index(x)[y].plot(kind="pie", autopct='%1.1f%%')
            elif chart_type == "bar":
                sns.barplot(x=x, y=y, hue=hue, data=df)
            elif chart_type == "hbar":
                sns.barplot(y=x, x=y, hue=hue, data=df, orient="h")
            elif chart_type == "line":
                sns.lineplot(x=x, y=y, data=df)
            elif chart_type == "hist":
                sns.histplot(df[y], bins=10, kde=True)
            elif chart_type == "scatter":
                sns.scatterplot(x=x, y=y, data=df)
            plt.title(title)
            plt.tight_layout()
            filepath = os.path.join("charts", filename)
            plt.savefig(filepath)
            plt.close()
            print(f"[OK] Сохранён график: {filepath}, строк данных: {len(df)}")

        graph_queries = [
            (
                "Pie chart: распределение игр по жанрам",
                """
                SELECT s.genres, COUNT(*) as game_count
                FROM steam s
                JOIN steam_description_data d ON s.appid = d.steam_appid
                GROUP BY s.genres
                ORDER BY game_count DESC
                LIMIT 5;
                """,
                "pie", "genres", "game_count", None, "pie_genres.png"
            ),
            (
                "Bar chart: средняя цена по издателям",
                """
                SELECT s.publisher, AVG(s.price) as avg_price
                FROM steam s
                JOIN steam_description_data d ON s.appid = d.steam_appid
                GROUP BY s.publisher
                HAVING COUNT(*) > 5
                ORDER BY avg_price DESC
                LIMIT 10;
                """,
                "bar", "publisher", "avg_price", None, "bar_publishers.png"
            ),
            (
                "Horizontal bar chart: топ разработчиков по играм",
                """
                SELECT s.developer, COUNT(*) as game_count
                FROM steam s
                JOIN steamspy_tag_data t ON s.appid = t.appid
                GROUP BY s.developer
                ORDER BY game_count DESC
                LIMIT 10;
                """,
                "hbar", "developer", "game_count", None, "hbar_devs.png"
            ),
            (
                "Line chart: динамика выхода игр по годам",
                """
                SELECT YEAR(s.release_date) as year, COUNT(*) as game_count
                FROM steam s
                JOIN steam_media_data m ON s.appid = m.steam_appid
                GROUP BY year
                ORDER BY year;
                """,
                "line", "year", "game_count", None, "line_years.png"
            ),
            (
                "Histogram: распределение цен на игры",
                """
                SELECT s.price
                FROM steam s
                JOIN steamspy_tag_data t ON s.appid = t.appid
                WHERE s.price > 0;
                """,
                "hist", None, "price", None, "hist_prices.png"
            ),
            (
                "Scatter plot: цена vs положительные отзывы",
                """
                SELECT s.price, s.positive_ratings
                FROM steam s
                JOIN steamspy_tag_data t ON s.appid = t.appid
                WHERE s.price > 0;
                """,
                "scatter", "price", "positive_ratings", None, "scatter_price_ratings.png"
            ),
        ]

        for title, query, chart_type, x, y, hue, filename in graph_queries:
            df = pd.read_sql(query, connection)
            make_chart(df, chart_type, title, x, y, hue, filename)

        print("\n=== Интерактивный график: Отзывы по годам (1999–2023) ===")
        time_query = """
            SELECT 
                YEAR(s.release_date) AS year,
                AVG(s.positive_ratings) AS avg_positive,
                AVG(s.negative_ratings) AS avg_negative
            FROM steam s
            WHERE s.release_date IS NOT NULL
            AND YEAR(s.release_date) BETWEEN 1999 AND 2023
            GROUP BY year
            ORDER BY year;
        """
        df_time = pd.read_sql(time_query, connection)

        df_melted = df_time.melt(
            id_vars=["year"],
            value_vars=["avg_positive", "avg_negative"],
            var_name="Метрика",
            value_name="Значение"
        )

        fig = px.bar(
            df_melted,
            x="Метрика",
            y="Значение",
            color="Метрика",
            animation_frame="year",
            title="Позитивные и негативные отзывы по годам (с ползунком времени)",
            labels={"Значение": "Среднее количество отзывов"},
            color_discrete_map={
                "avg_positive": "blue",   
                "avg_negative": "orange"   
            }
        )

        fig.update_layout(
            xaxis={'categoryorder': 'total descending'}
        )
        fig.show()

        print("\n=== Экспорт в Excel ===")
        export_query = """
            SELECT s.name, s.price, s.positive_ratings, s.negative_ratings, s.developer, s.publisher
            FROM steam s
            JOIN steamspy_tag_data t ON s.appid = t.appid
            LIMIT 200;
        """
        df_export = pd.read_sql(export_query, connection)

        filepath = os.path.join("exports", "games_report.xlsx")
        df_export.to_excel(filepath, index=False, engine="openpyxl")

        wb = load_workbook(filepath)
        ws = wb.active

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"

        from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
        from openpyxl.styles import PatternFill

        price_col = "B2:B{}".format(ws.max_row)
        color_scale_rule = ColorScaleRule(start_type="min", start_color="00FF00", end_type="max", end_color="FF0000")
        ws.conditional_formatting.add(price_col, color_scale_rule)

        pos_range = f"C2:C{ws.max_row}"
        blue_gradient = ColorScaleRule(start_type="num", start_value=0, start_color="FFFFFF", end_type="num", end_value=50000, end_color="0000FF")
        ws.conditional_formatting.add(pos_range, blue_gradient)

        neg_range = f"D2:D{ws.max_row}"
        orange_gradient = ColorScaleRule(start_type="num", start_value=0, start_color="FFFFFF", end_type="num", end_value=30000, end_color="FFA500")
        ws.conditional_formatting.add(neg_range, orange_gradient)

        wb.save(filepath)

        print(f"[OK] Создан Excel отчёт: {filepath}, строк: {len(df_export)}")

except Error as e:
    print("Ошибка:", e)
